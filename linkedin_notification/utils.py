# utils.py
from selenium import webdriver
driver = webdriver.Chrome()
from selenium import webdriver
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook
import datetime


def login_to_linkedin(username, encrypted_password):
    options = Options()
    options.add_argument("--headless")  
    driver = webdriver.Chrome(options=options)
    url = 'https://www.linkedin.com/login'

    driver.get(url)
    time.sleep(3)

    username_input = driver.find_element("xpath","//input[@name = 'session_key']")
    username_input.send_keys(username)

    password_input = driver.find_element("xpath","//input[@name = 'session_password']")
    password_input.send_keys(encrypted_password)
    time.sleep(2)
    password_input.send_keys(Keys.ENTER)

    return driver


def retrieve_unread_messages(driver):
    driver.get("https://www.linkedin.com/messaging")

    unread_messages_element = driver.find_element("xpath","//*[@id="global-nav"]/div/nav/ul/li[5]/a")
    unread_messages = int(unread_messages_element.text)

    return unread_messages


def retrieve_unread_notifications(driver):
    driver.get("https://www.linkedin.com/notifications/")

    unread_notifications_element = driver.find_element("xpath","//*[@id="global-nav"]/div/nav/ul/li[4]/a")
    unread_notifications = int(unread_notifications_element.text)

    return unread_notifications


def store_data_to_excel(data):
    current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    data.append(current_time)

    wb = load_workbook("linkedin_data.xlsx")
    ws = wb.active

    ws.append(data)
    wb.save("linkedin_data.xlsx")
    wb.close()


def compare_data_with_prev(current_data):
    wb = load_workbook("linkedin_data.xlsx")
    ws = wb.active

    previous_row = ws.max_row - 1
    previous_data = [ws.cell(previous_row, i).value for i in range(1, 5)]

    comparison_result = {
        "unread_messages_change": current_data[0] - previous_data[0],
        "unread_notifications_change": current_data[1] - previous_data[1]
    }

    wb.close()

    return comparison_result
